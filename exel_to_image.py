from openpyxl import load_workbook
from PIL import Image, ImageDraw
import cv2 as cv
import numpy as np
import matplotlib.pyplot as plt
import os

def create_image_from_excel(file_path, sheet_name, top_left, bottom_right, output_file):

    # 엑셀 파일 불러오기
    wb = load_workbook(file_path)

    # 시트 선택
    ws = wb[sheet_name]

    # 캡처할 범위 지정
    range_to_check = ws[top_left:bottom_right]

    # 이미지 크기 설정
    img_width = len(range_to_check[0])
    img_height = len(range_to_check)
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)


    # 셀 값이 17.0 이하인 경우에만 점을 그림
    for row_idx, row in enumerate(range_to_check):
        for col_idx, cell in enumerate(row):
            if cell.value is not None and\
            isinstance(cell.value, (int, float)) and\
            16.0 <= cell.value <= 18.0:
                draw.point((col_idx, row_idx), fill='black')

    # 이미지 저장
    img.save(output_file)

    return img


# 사용 예시
#excel_file = "./exel/I18700.xlsx"
#sheet_name = "Sheet1"
#top_left_cell = "EJ163"
#bottom_right_cell = "ZL1391"
#output_image = "exel_to_image.png"
#img = create_image_from_excel(excel_file, sheet_name, top_left_cell, bottom_right_cell, output_image)



def process_excel_files_in_folder(folder_path, sheet_name, top_left, bottom_right, output_folder):

    # 폴더 내의 모든 파일 가져오기
    for file_name in os.listdir(folder_path):

        if file_name.endswith('.xlsx'):  # xlsx 파일만 처리

            # 파일의 전체 경로 구성
            file_path = os.path.join(folder_path, file_name)

            # 캡처된 이미지를 저장할 파일 이름 구성
            output_file = os.path.join(output_folder, file_name[:-5] + '.png')

            # 엑셀 파일 처리
            create_image_from_excel(file_path, sheet_name, top_left, bottom_right, output_file)



# 사용 예시
folder_path = "./exel"  # 엑셀 파일이 있는 폴더 경로
sheet_name = "Sheet1"
top_left_cell = "EJ163" # 고정
bottom_right_cell = "ZL1391" # 고정
output_folder = "./images_from_excel"  # 캡처된 이미지를 저장할 폴더 경로

# 폴더 내의 모든 엑셀 파일 처리
process_excel_files_in_folder(folder_path, sheet_name, top_left_cell, bottom_right_cell, output_folder)