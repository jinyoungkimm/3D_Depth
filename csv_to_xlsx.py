from openpyxl import load_workbook
from PIL import ImageGrab
import os
import pandas as pd
import csv
import cv2 as cv
import numpy as np

# convert_csv_to_xlsx() 함수는 초기에 단 1번만 실행하면 된다.
# 1] cvs 폴더 내의 모든 csv 파일을 개선 후 제거
# 2] 새로운 cvs 파일들을 cvs 폴더 내에 똑같은 이름으로 저장
# 3] cvs 폴더 내의 새로운 cvs 파일들을 xlsx로 전환하여, exel 폴더에 저장
def convert_csv_to_xlsx(folder_path):

    # 폴더 내의 모든 파일 가져오기
    for file_name in os.listdir(folder_path):

        if file_name.endswith('.csv'):  # CSV 파일인 경우에만 처리

            # CSV 파일의 전체 경로 구성
            csv_file_path = os.path.join(folder_path, file_name)

            # XLSX 파일의 이름 구성 (확장자를 .xlsx로 변경)
            xlsx_file_name = os.path.splitext(file_name)[0] + '.xlsx'

            # XLSX 파일의 전체 경로 구성
            xlsx_file_path = os.path.join("./exel", xlsx_file_name)

            # CSV 파일 열기 및 데이터 읽기
            with open(csv_file_path, 'rt') as file:
                reader = csv.reader(file)
                csv_list = [row for row in reader]

            # 기존 파일 삭제
            os.remove(csv_file_path)

            # DataFrame 생성
            df = pd.DataFrame(csv_list)

            # 새로운 CSV 파일로 저장
            new_file_path = os.path.join(folder_path, f"{file_name}")
            df.to_csv(new_file_path, index=False, header=False) # 개선 전 csv 파일 -> 개선 후 csv 파일("new_~~~.csv")

            # 새롭운 CSV 파일을 읽어와서 DataFrame으로 변환
            df = pd.read_csv(new_file_path, low_memory=False) # 반환 형태 : DataFrame

            # DataFrame을 XLSX 파일로 저장
            df.to_excel(xlsx_file_path, index=False)




# 사용 예시
#folder_path = "./csv"  # CSV 파일이 있는 폴더 경로

# 폴더 내의 모든 CSV 파일을 XLSX로 변환
#convert_csv_to_xlsx(folder_path)



def capture_excel_range(file_path, sheet_name, top_left, bottom_right, output_file):


    # 엑셀 파일 불러오기
    wb = load_workbook(file_path)

    # 시트 선택
    ws = wb[sheet_name]

    # 캡처할 범위 지정
    top_left_cell = ws[top_left]
    bottom_right_cell = ws[bottom_right]

    # 범위의 셀 좌표 가져오기
    top = top_left_cell.row
    left = top_left_cell.column
    bottom = bottom_right_cell.row
    right = bottom_right_cell.column

    # 캡처할 영역의 좌표 계산
    top_left_pixel = ws[top_left].coordinate
    bottom_right_pixel = ws[bottom_right].coordinate

    # 엑셀의 좌표를 픽셀 좌표로 변환
    top_left_pixel = ws[top_left_pixel].row
    bottom_right_pixel = ws[bottom_right_pixel].column

    # 이미지 캡처
    img = ImageGrab.grab(bbox=(top_left_pixel, left, bottom_right_pixel, right))
    img = np.array(img)
    print(img)
    print(output_file)
    imgs = cv.cvtColor(img,cv.COLOR_RGB2BGR)
    cv.imshow("adf",imgs)
    cv.waitKey(0)

    # 이미지 저장
    #img.save(output_file)
    #cv.imwrite(img,output_file)


folder_path = "./H27178.xlsx"  # 엑셀 파일이 있는 폴더 경로
sheet_name = "Sheet1"
top_left_cell = "IZ166" # IZ166
bottom_right_cell = "WH1351" # WH1351
output_folder = "./output_images"  # 캡처된 이미지를 저장할 폴더 경로

capture_excel_range(folder_path, sheet_name, top_left_cell, bottom_right_cell, output_folder)


def process_excel_files_in_folder(folder_path, sheet_name, top_left, bottom_right, output_folder):

    # 폴더 내의 모든 파일 가져오기
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):  # xlsx 파일만 처리

            print("file name",file_name)

            # 파일의 전체 경로 구성
            file_path = os.path.join(folder_path, file_name)

            # 캡처된 이미지를 저장할 파일 이름 구성
            output_file = os.path.join(output_folder, file_name[:-5] + '_captured_range.jpg')

            # 엑셀 파일 처리
            capture_excel_range(file_path, sheet_name, top_left, bottom_right, output_file)

"""""""""
# 사용 예시
folder_path = "./exel"  # 엑셀 파일이 있는 폴더 경로
sheet_name = "Sheet1"
top_left_cell = "IZ166" # IZ166
bottom_right_cell = "WH1351" # WH1351
output_folder = "./output_images"  # 캡처된 이미지를 저장할 폴더 경로

# 폴더 내의 모든 엑셀 파일 처리
process_excel_files_in_folder(folder_path, sheet_name, top_left_cell, bottom_right_cell, output_folder)

"""""""""

